#!/usr/bin/env python3
"""
compare.py
对比两个解码目录中同名图片的色彩差异，并输出 Excel 报告（多 sheet）。

默认目录：
- RawTherapee 输出: C:\pack\file\output\rawtherapee
- raw 输出       : C:\pack\file\output\raw

输出文件：
- RawTherapee\installer\compare_report.xlsx
"""

from __future__ import annotations

import argparse
import math
import shutil
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
from PIL import Image

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 openpyxl，请先安装: pip install openpyxl\n"
        f"原始错误: {exc}"
    )

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="比较 rawtherapee 与 raw 解码结果色彩差异")
    parser.add_argument(
        "--rt-dir",
        default=r"C:\pack\file\output\rawtherapee",
        help="RawTherapee 输出目录",
    )
    parser.add_argument(
        "--raw-dir",
        default=r"C:\pack\file\output\raw",
        help="raw 输出目录",
    )
    parser.add_argument(
        "--input-dir",
        default=r"C:\pack\file\input",
        help="原始输入目录",
    )
    parser.add_argument(
        "--diff-copy-dir",
        default=r"C:\pack\file\差异",
        help="差异图片拷贝目录",
    )
    parser.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent / "compare_report.xlsx"),
        help="输出 Excel 文件路径",
    )
    parser.add_argument(
        "--luma-threshold",
        type=float,
        default=8.0,
        help="luma 均值差阈值（大于判为差异大）",
    )
    parser.add_argument(
        "--mae-threshold",
        type=float,
        default=6.0,
        help="RGB 平均 MAE 阈值（大于判为差异大）",
    )
    return parser.parse_args()


def list_images(root: Path) -> Dict[str, Path]:
    mapping: Dict[str, Path] = {}
    for p in root.rglob("*"):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
            rel = str(p.relative_to(root)).replace("\\", "/").lower()
            mapping[rel] = p
    return mapping


def load_rgb(path: Path) -> np.ndarray:
    with Image.open(path) as img:
        return np.asarray(img.convert("RGB"), dtype=np.float32)


def compute_metrics(rt_img: np.ndarray, raw_img: np.ndarray) -> Dict[str, float]:
    if rt_img.shape != raw_img.shape:
        h = min(rt_img.shape[0], raw_img.shape[0])
        w = min(rt_img.shape[1], raw_img.shape[1])
        rt_img = rt_img[:h, :w, :]
        raw_img = raw_img[:h, :w, :]

    diff = rt_img - raw_img
    abs_diff = np.abs(diff)

    rt_mean = rt_img.mean(axis=(0, 1))
    raw_mean = raw_img.mean(axis=(0, 1))

    rt_luma = 0.2126 * rt_img[:, :, 0] + 0.7152 * rt_img[:, :, 1] + 0.0722 * rt_img[:, :, 2]
    raw_luma = 0.2126 * raw_img[:, :, 0] + 0.7152 * raw_img[:, :, 1] + 0.0722 * raw_img[:, :, 2]

    mse = float(np.mean(diff * diff))
    psnr = 99.0 if mse <= 1e-12 else 20.0 * math.log10(255.0 / math.sqrt(mse))

    return {
        "rt_r_mean": float(rt_mean[0]),
        "rt_g_mean": float(rt_mean[1]),
        "rt_b_mean": float(rt_mean[2]),
        "raw_r_mean": float(raw_mean[0]),
        "raw_g_mean": float(raw_mean[1]),
        "raw_b_mean": float(raw_mean[2]),
        "delta_r_mean": float(rt_mean[0] - raw_mean[0]),
        "delta_g_mean": float(rt_mean[1] - raw_mean[1]),
        "delta_b_mean": float(rt_mean[2] - raw_mean[2]),
        "rt_luma_mean": float(rt_luma.mean()),
        "raw_luma_mean": float(raw_luma.mean()),
        "delta_luma_mean": float(rt_luma.mean() - raw_luma.mean()),
        "mae_r": float(abs_diff[:, :, 0].mean()),
        "mae_g": float(abs_diff[:, :, 1].mean()),
        "mae_b": float(abs_diff[:, :, 2].mean()),
        "mae_rgb_mean": float(abs_diff.mean()),
        "psnr": float(psnr),
    }


def classify(metrics: Dict[str, float], luma_th: float, mae_th: float) -> str:
    if abs(metrics["delta_luma_mean"]) > luma_th or metrics["mae_rgb_mean"] > mae_th:
        return "有"
    return "无"


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = name.replace("\\", "_").replace("/", "_")
    cleaned = cleaned.replace(":", "_").replace("*", "_").replace("?", "_")
    cleaned = cleaned.replace("[", "(").replace("]", ")")
    if not cleaned:
        cleaned = "root"
    base = cleaned[:31]
    candidate = base
    idx = 1
    while candidate in used:
        suffix = f"_{idx}"
        candidate = (base[: 31 - len(suffix)] + suffix)
        idx += 1
    used.add(candidate)
    return candidate


def write_excel(rows: List[Dict[str, object]], out_file: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "relative_path",
        "rt_path",
        "raw_path",
        "rt_r_mean", "rt_g_mean", "rt_b_mean",
        "raw_r_mean", "raw_g_mean", "raw_b_mean",
        "delta_r_mean", "delta_g_mean", "delta_b_mean",
        "rt_luma_mean", "raw_luma_mean", "delta_luma_mean",
        "mae_r", "mae_g", "mae_b", "mae_rgb_mean",
        "psnr",
        "差异大(有/无)",
    ]

    by_group: Dict[str, List[Dict[str, object]]] = {}
    for r in rows:
        rel = str(r["relative_path"])
        parent = str(Path(rel).parent).replace("\\", "/")
        if parent == ".":
            parent = "root"
        by_group.setdefault(parent, []).append(r)

    used_names: set[str] = set()

    summary = wb.create_sheet(title="summary")
    summary.append([
        "group", "images", "差异大数量", "差异大占比(%)",
        "avg_abs_delta_luma", "avg_mae_rgb",
    ])

    for group, group_rows in sorted(by_group.items()):
        ws = wb.create_sheet(title=safe_sheet_name(group, used_names))
        ws.append(headers)
        diff_count = 0
        abs_luma = []
        mae_vals = []

        for r in group_rows:
            if r["差异大(有/无)"] == "有":
                diff_count += 1
            abs_luma.append(abs(float(r["delta_luma_mean"])))
            mae_vals.append(float(r["mae_rgb_mean"]))
            ws.append([r.get(h, "") for h in headers])

        total = len(group_rows)
        ratio = (diff_count * 100.0 / total) if total else 0.0
        summary.append([
            group,
            total,
            diff_count,
            round(ratio, 2),
            round(float(np.mean(abs_luma)) if abs_luma else 0.0, 4),
            round(float(np.mean(mae_vals)) if mae_vals else 0.0, 4),
        ])

    wb.save(out_file)


def load_diff_rows_from_report(report_file: Path) -> List[Dict[str, object]]:
    wb = load_workbook(report_file, data_only=True)
    rows: List[Dict[str, object]] = []

    for ws in wb.worksheets:
        if ws.title.lower() == "summary":
            continue

        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            continue

        headers = [str(h).strip() if h is not None else "" for h in header_cells]
        try:
            rel_idx = headers.index("relative_path")
            flag_idx = headers.index("差异大(有/无)")
        except ValueError:
            continue

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            rel = r[rel_idx] if rel_idx < len(r) else None
            flag = r[flag_idx] if flag_idx < len(r) else None
            if not rel:
                continue
            rows.append({
                "relative_path": str(rel).replace("\\", "/"),
                "差异大(有/无)": str(flag) if flag is not None else "",
            })

    return rows


def copy_diff_images(
    rows: List[Dict[str, object]],
    input_dir: Path,
    rt_dir: Path,
    raw_dir: Path,
    diff_copy_dir: Path,
) -> int:
    if diff_copy_dir.exists():
        shutil.rmtree(diff_copy_dir)

    copied = 0
    for row in rows:
        if row.get("差异大(有/无)") != "有":
            continue

        rel = str(row.get("relative_path", "")).replace("\\", "/")
        if not rel:
            continue

        stem = Path(rel).stem
        parent = Path(rel).parent

        input_src = input_dir / parent / f"{stem}.CR2"
        if not input_src.exists():
            input_src = input_dir / parent / Path(rel).name

        rt_src = rt_dir / rel
        raw_src = raw_dir / rel

        targets = [
            (input_src, diff_copy_dir / "input" / parent / input_src.name),
            (rt_src, diff_copy_dir / "rawtherapee" / parent / Path(rel).name),
            (raw_src, diff_copy_dir / "raw" / parent / Path(rel).name),
        ]

        for src, dst in targets:
            if src.exists():
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)
        copied += 1

    return copied


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    rt_dir = Path(args.rt_dir)
    raw_dir = Path(args.raw_dir)
    diff_copy_dir = Path(args.diff_copy_dir)
    out_file = Path(args.out)

    if not input_dir.is_dir():
        raise SystemExit(f"input 目录不存在: {input_dir}")
    if not rt_dir.is_dir():
        raise SystemExit(f"RawTherapee 输出目录不存在: {rt_dir}")
    if not raw_dir.is_dir():
        raise SystemExit(f"raw 输出目录不存在: {raw_dir}")

    rows: List[Dict[str, object]] = []

    if out_file.is_file():
        print(f"检测到已有报告，直接按报告拷贝差异图片: {out_file}")
        rows = load_diff_rows_from_report(out_file)
        if not rows:
            raise SystemExit(f"报告中未解析到有效记录: {out_file}")
    else:
        rt_map = list_images(rt_dir)
        raw_map = list_images(raw_dir)
        common = sorted(set(rt_map.keys()) & set(raw_map.keys()))

        if not common:
            raise SystemExit("两个目录没有可匹配的同名图片，请先完成两边解码输出。")

        total = len(common)
        for idx, rel in enumerate(common, start=1):
            rt_path = rt_map[rel]
            raw_path = raw_map[rel]

            print(f"[compare] 进度: {idx}/{total}")
            print(f"[compare] 当前: {rel}")

            try:
                rt_img = load_rgb(rt_path)
                raw_img = load_rgb(raw_path)
                metrics = compute_metrics(rt_img, raw_img)
                flag = classify(metrics, args.luma_threshold, args.mae_threshold)
                row = {
                    "relative_path": rel,
                    "rt_path": str(rt_path),
                    "raw_path": str(raw_path),
                    **{k: round(v, 4) for k, v in metrics.items()},
                    "差异大(有/无)": flag,
                }
            except Exception as exc:
                row = {
                    "relative_path": rel,
                    "rt_path": str(rt_path),
                    "raw_path": str(raw_path),
                    "差异大(有/无)": "有",
                    "psnr": "ERROR",
                    "mae_rgb_mean": "ERROR",
                    "delta_luma_mean": f"ERROR: {exc}",
                }
            rows.append(row)

        out_file.parent.mkdir(parents=True, exist_ok=True)
        write_excel(rows, out_file)

    diff_count = sum(1 for r in rows if r.get("差异大(有/无)") == "有")
    copied_groups = copy_diff_images(rows, input_dir, rt_dir, raw_dir, diff_copy_dir)

    print(f"完成对比: {len(rows)} 张")
    print(f"差异大(有): {diff_count} 张")
    print(f"报告输出: {out_file}")
    print(f"差异图片已拷贝: {copied_groups} 组")
    print(f"差异目录: {diff_copy_dir}")


if __name__ == "__main__":
    main()
